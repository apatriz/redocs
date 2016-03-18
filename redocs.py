#!/usr/bin/env python
#AUTHOR: Alec Patrizio

""" This python script reads a spreadsheet column containing file names into a list,
    and compares it to a list of all files in a specific directory. The purpose is to identify files
    that should be preserved, delete the rest, execute a cleanup of the directory structure to eliminate
    empty directories and produce a deliverable consisting of only the needed files.
    It also creates a new workbook.xlsx , containing only the entries for the preserved files."""


import os
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import column_index_from_string

class DocSorter(object):


        def __init__(self,filename,lpdir,fn_column,ident_column,keep_val,fold_col,
                     comment_col,col_header):
                
                self.filename = filename
                self.lpdir = lpdir
                self.fn_column = fn_column
                self.ident_column = ident_column
                self.keep_val = keep_val
                self.fold_col = fold_col
                self.comment_col = comment_col
                self.col_header = col_header
                self.wb = load_workbook(self.filename, use_iterators= True)
                self.ws = self.wb.active
                
        
        

        # generator to get all files (full paths) in a directory (recursive).
        def get_docs(self):
                for path,dirs,files in os.walk(self.lpdir):
                        for fn in files:
                                yield os.path.join(path,fn)

                                
        # read values from a spreadsheet into a list, based on keep value
        def tokeep(self):
                keeplist = []
                for row in self.ws.iter_rows():
                        if row[self.ident_column].value and(row[self.ident_column].value).strip() == self.keep_val:
                                keeplist.append(row[self.fn_column].value)
                return keeplist


        # compare the two lists and deletes files that are not in 'tokeep' list
        def del_unwanted(self):
                for i in self.get_docs():
                        if os.path.splitext(os.path.split(i)[1])[0] not in self.tokeep():
                                # print i, "will be deleted"
                                os.remove(i)
                                print i, "has been deleted."
                print "All unwanted files deleted."


        # recursively yield empty directory paths (deleting files may leave empty directories
        # that should be cleaned up)
        def get_empty_dir(self):
                for path,dirs,files in os.walk(self.lpdir):
                        if not dirs and not files:
                                yield path

        # delete empty directories using generated list from GetEmptyDir.
        # will run until no empty directories are found.
        def cleanup(self):
                while True:
                        empty_dir = list(self.get_empty_dir())
                        for i in empty_dir:
                                os.rmdir(i)
                                print i, 'is empty and has been deleted.'
                        if empty_dir == []:
                                print 'All empty directories deleted.'
                                break

        # return a list of data entries for the preserved files. Each entry in the list is a tuple
        # containing the needed data for each file (i.e.(folder, filename, comments)).
        # the first entry will always be the column header labels
        def create_entry_list(self):
                entrylist = [(self.col_header)]
                prevfol = ''
                for row in self.ws.iter_rows():
                        folder = row[self.fold_col].value
                        fn = row[self.fn_column].value
                        comment = row[self.comment_col].value
                        if not folder:
                                if not fn:
                                        continue
                                folder = prevfol
                        else:
                                prevfol = folder
                        if row[self.ident_column].value == self.keep_val:
                                entry = folder,fn,comment
                                entrylist.append(entry)
                return entrylist 

        # create new workbook.xlsx named "Returned" inside the directory containing the preserved files,
        # which contains the data entries for the preserved files (from create_entry_list)
        def create_new_xl(self):
                nb = Workbook(write_only=True)
                ns = nb.create_sheet()
                output = os.path.join(self.lpdir,'Returned.xlsx')
                entrylist = self.create_entry_list()
                for entry in entrylist:
                        ns.append([value for value in entry])
                nb.save(output)
                print "Saved new spreadsheet containing preserved file entries to {0}".format(output)

        # executes the main processes 
        def run(self):
                self.del_unwanted()
                self.cleanup()
                self.create_new_xl()
                print "Doc sorter has completed."
                


# define the main function
def main():

# define decorator to get 0 based index from spreadsheet string index 
    def convertindex(func):
        def minus_one(index_string):
            return func(index_string) - 1
        return minus_one
        
    # decorate column_index_from_string
    get_index = convertindex(column_index_from_string)

##  filename = os.path.abspath(raw_input("Enter full path to spreadsheet file (include file extension): "))
    filename = os.path.abspath('C:\Users\patrizio\Projects\Macomb_ROW\Deliverables\Lake_Returned_Docs.xlsx')

##  lpdir = os.path.abspath(raw_input("Enter full path of directory containing the documents: "))
    lpdir = os.path.abspath('C:\Users\patrizio\Projects\Macomb_ROW\Deliverables\Lake(Southeast)_Working')

##  fn_column = get_index(raw_input('''Enter the spreadsheet column which lists the file names
##                                        to be preserved (i.e. enter 'A' for column A, 'B' for column 'B',etc.): '''))

    # using column_index_from_string() returns index numbers that begin at '1', therefore we must subtract 1 because the spreadsheet cells are read
    # and written as an iterable list using the optimized reader/writer in openpyxl, so column indexes in the iterables will start at '0'
    fn_column = get_index('B')

##  ident_column = get_index(raw_input('''Enter the spreadsheet column which lists the values
##                                           for identifying which files should be preserved: '''))
    ident_column = get_index('C')

##  keep_val = raw_input("Enter the value that identifies which files should be preserved: ")
    keep_val = 'Not Drawn'

##  fold_col = get_index(raw_input("Enter the spreadsheet column which lists the file locations: "))
    fold_col = get_index('A')

##  comment_col = get_index(raw_input("Enter the spreadsheet column which lists the comments: "))
    comment_col = get_index('D')

##  col_header = tuple(raw_input('''Enter the column labels for the output spreadsheet
##                                     (i.e. 'Folder,Liber Page,Comments' for columns 1, 2 and 3 labels respectively): ''').split(','))
    
    col_header = ('Folder','Liber Page','Comments')

    sorter = DocSorter(filename=filename,lpdir=lpdir,fn_column=fn_column,ident_column=ident_column,keep_val=keep_val,
                           fold_col = fold_col,comment_col = comment_col,col_header=col_header)
    sorter.run()
                  
##
##  # load workbook and active sheet
##  global wb
##  global ws
##  wb = load_workbook(filename, use_iterators= True)
##  ws = wb.active
##
##  
##  params = {'filename':filename,'lpdir':lpdir,'fn_column':fn_column,
##                  'ident_column':ident_column,'keep_val':keep_val,'fold_col':fold_col,
##                  'comment_col':comment_col,'col_header':col_header}
##  
##  # execute function calls
##  del_unwanted(**params)
##  cleanup(**params)
##  create_new_xl(**params)
##  
##  del_unwanted(lpdir,filename)
##  cleanup(lpdir)
##  create_new_xl(filename)

# call the main function
if __name__ == "__main__":
    main()
    
