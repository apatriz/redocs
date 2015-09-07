#!/usr/bin/env python
##TODO: Add functionality to create new spreadsheet containing only
## entries for files that are preserved


import os
from openpyxl import load_workbook
filename = raw_input("Enter full path to spreadsheet file (include file extension): ")
#'C:\\Macomb ROW\\Returned Docs Delivery Script\\test\\Armada_Not_Drawn.xlsx'
lpdir = raw_input("Enter full path of directory containing the documents: ")
#'C:/Macomb ROW/Returned Docs Delivery Script/test/Armada Liber Page Docs'
fn_column = raw_input("Enter the spreadsheet column which lists the file names to be preserved (i.e. Enter '0' for column 'A', '1' for column 'B',etc.): ")
ident_column = raw_input("Enter the spreadsheet column which lists the values for identifying which files should be preserved: ")
keep_val = raw_input("Enter the value that identifies which files should be preserved: ")

wb = load_workbook(filename, use_iterators= True)
ws = wb.get_active_sheet()


#Define generator to get all files (full paths) in a directory (recursive).
def get_docs(lpdir):
    for path,dirs,files in os.walk(lpdir):
        for fn in files:
            yield os.path.join(path,fn)

#Define function that reads values from a spreadsheet into a list, based on user input
#TODO: transform user input for column choice ("A","B",etc.) into corresponding integer val
def tokeep(filename):
    keeplist = []
    for row in ws.iter_rows():
        if row[int(ident_column)].value == keep_val:
            keeplist.append(row[int(fn_column)].value)
    return keeplist


 # Define function that compares the two lists and deletes files that are not in 'tokeep' list
def comp_lists(lpdir,filename):
    for i in get_docs(lpdir):
        if os.path.splitext(os.path.split(i)[1])[0] not in tokeep(filename):
            # print i, "will be deleted"
            os.remove(i)
            print i, "has been deleted."


# Define generator object to recursively yield empty directory paths (deleting files may leave empty directories
# that should be cleaned up)
def get_empty_dir(lpdir):
    for path,dirs,files in os.walk(lpdir):
        if not dirs and not files:
            yield path

# Define function to delete empty directories using generated list from GetEmptyDir.
# Will run until no empty directories are found.
def cleanup(lpdir):
    while True:
        empty_dir = list(get_empty_dir(lpdir))
        for i in empty_dir:
            os.rmdir(i)
            print i, 'is empty and has been deleted.'
        if empty_dir == []:
            print 'All empty directories have been deleted.'
            break


##print list(get_docs(lpdir))
# print tokeep(filename)
print comp_lists(lpdir,filename)
print cleanup(lpdir)
